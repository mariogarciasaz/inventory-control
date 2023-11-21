from typing import Any
from django.db.models.query import QuerySet
from django.http import HttpResponse
from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView
from clients.models import Client
from django.db.models import Q
from clients.forms import ClientForm
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from django.contrib.auth.mixins import LoginRequiredMixin


# Create your views here.



#CLIENTS VIEWS
##############


class ClientList(LoginRequiredMixin, ListView):
    model = Client
    template_name = 'client_list.html'
    context_object_name = 'clients'


    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(fullname__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(phone__icontains=search_query) |
                Q(address__icontains=search_query)
                )
        return queryset
    

    def get_context_data(self, **kwargs):
        context =  super().get_context_data(**kwargs)

        all_clients = Client.objects.all()
        context['all_clients'] = len(all_clients)

        return context

    

    def report_clients(request):

        queryset = Client.objects.all()
        workbook = Workbook()
        sheet = workbook.active
        
        sheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=4)
        merged_cell = sheet.cell(row=1, column=1)
        merged_cell.value = "Reporte de todos los clientes"
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.font = Font(bold=True, size=20, color='FFFFFF')
        fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        merged_cell.fill = fill
        sheet['A1'].font = Font(bold=True, size=20, color='FFFFFF', italic=True)
        row_dimension = sheet.row_dimensions[5]
        row_dimension.height = 30
        
        all_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'O', 'P']
        for column in all_columns:
            sheet.column_dimensions[column].width = 50
        
        headers = ['Nombre Completo', 'Email', 'Telefono', 'Direccion']
        sheet.append([''] * len(headers))
        sheet.append([''] * len(headers))
        sheet.append([''] * len(headers))
        sheet.append(headers)
        
        header_border = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )
        
        for header_cell in sheet[5]:
            header_cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
            header_cell.font = Font(bold=True, color='FFFFFF', size=14)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = header_border
        
        row_index = 3
        for i in queryset:
            row = [
                i.fullname,
                i.email,
                i.phone,
                i.address
            ]
            sheet.append(row)
            row_index += 1
        
        data_border = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )
        
        for row in sheet.iter_rows(min_row=6, max_row=row_index+2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = data_border

        filename = "Reporte de Clientes.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response





class AddClient(LoginRequiredMixin, CreateView):
    model = Client
    template_name = 'add_client.html'
    form_class = ClientForm
    success_url = reverse_lazy('clients:Clients')

    def form_valid(self, form):
        if self.request.method == 'POST':
            form.instance.fullname = self.request.POST.get('fullname')
            form.instance.email = self.request.POST.get('email')
            form.instance.phone = self.request.POST.get('phone')
            form.instance.address = self.request.POST.get('address')
        return super().form_valid(form)
    

class UpdateClient(LoginRequiredMixin, UpdateView):
    model = Client
    template_name = 'update_client.html'
    form_class = ClientForm
    success_url = reverse_lazy('clients:Clients')

    def form_valid(self, form):
        form.instance.fullname = form.instance.fullname
        return super().form_valid(form)
    

class DeleteClient(LoginRequiredMixin, DeleteView):
    model = Client
    template_name = 'delete_client.html'
    success_url = reverse_lazy('clients:Clients')


