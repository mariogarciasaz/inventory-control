from decimal import Decimal
from typing import Any
from django.db import models
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from clientsdata.models import ClientData
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from clientsdata.forms import ClientDataForm
from django.db.models import Q
from clients.models import Client
from products.models import Product
from django.forms import formset_factory
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.db.models.functions import ExtractYear

# Create your views here.


@login_required(login_url='website:login')
def client_data(request, pk):

    #GET CLIENT DATA
    client = Client.objects.get(id=pk)
    
    client_context = {
        'id': client.pk,
        'fullname': client.fullname,
        'email': client.email,
        'phone': client.phone,
        'address': client.address,
        'client_pk': pk
    }

    #GET CLIENTDATA DATA

    #TOTAL CLIENT LOANS
    client_data_list = ClientData.objects.filter(client_id=pk)
    client_data_list_count = len(client_data_list)




    #TOTAL CLIENT DEBT
    client_data_amount = []
    for cd in client_data_list:
        if cd.paid == True:
            client_data_amount.append(cd.total_deuda)

    client_data_amount_total = sum(client_data_amount)



    #TOTAL LOANS WITHOUT PAY
    client_data_without_pay = []
    for nopay in client_data_list:
        if nopay.paid == False:
            client_data_without_pay.append(nopay)

    total_without_pay = len(client_data_without_pay)



    #TOTAL DEBT WITHOUT PAY
    client_data_debt_without_pay = []
    for nopaydebt in client_data_list:
        if nopaydebt.paid == False:
            client_data_debt_without_pay.append(nopaydebt.total_deuda)

    total_debt_without_pay = sum(client_data_debt_without_pay)

    search_query = request.GET.get('search')
    if search_query:
        client_data_list = client_data_list.filter(
            Q(product__name__icontains=search_query) |
            Q(cantidad_enviada__icontains=search_query) |
            Q(cantidad_vendida__icontains=search_query) |
            Q(cantidad_retirada__icontains=search_query) |
            Q(start_date__icontains=search_query) |
            Q(paid__icontains=search_query) |
            Q(total_deuda__icontains=search_query)
        )
    year_filter = request.GET.get('start_date')
    if year_filter:
        client_data_list = client_data_list.filter(start_date__year=year_filter)
    all_years = ClientData.objects.filter(client_id=pk).values_list(ExtractYear('start_date'), flat=True)
    unique_years = list(set(all_years))

    


    return render(request, 'client_data_list.html', {
        'client_context': client_context,
        'client_data_list': client_data_list,
        'client_data_list_count': client_data_list_count,
        'client_data_amount_total': client_data_amount_total,
        'total_without_pay': total_without_pay,
        'total_debt_without_pay': total_debt_without_pay,
        'unique_years': unique_years
    })

def generate_excel_report_client(request, pk):

        selected_year = request.GET.get('start_date')
        if selected_year is not None and selected_year != '':
            try:
                # Intenta convertir selected_year a entero
                selected_year = int(selected_year)
                
            except ValueError:
                    print (selected_year)
                    pass
            
            queryset = ClientData.objects.filter(client_id=pk, start_date__year=selected_year)
    
        else:
            queryset = ClientData.objects.filter(client_id=pk)
    
    
        client_name = []
        for name in queryset:
            if name.client not in client_name:
                client_name.append(name.client.fullname)
        
        workbook = Workbook()
        sheet = workbook.active
        
        sheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=7)
        merged_cell = sheet.cell(row=1, column=1)
        merged_cell.value = f"Reporte de Consignas del Cliente {client_name[0]}"
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.font = Font(bold=True, size=20, color='FFFFFF')
        fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        merged_cell.fill = fill
        sheet['A1'].font = Font(bold=True, size=20, color='FFFFFF', italic=True)
        row_dimension = sheet.row_dimensions[5]
        row_dimension.height = 30
        
        all_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'O', 'P']
        for column in all_columns:
            sheet.column_dimensions[column].width = 20
        
        headers = ['Producto', 'Cantidad Enviada', 'Cantidad Vendida', 'Cantidad Retirada', 'Fecha', 'Total Deuda', 'Pagada']
        sheet.append([''] * len(headers))
        sheet.append([''] * len(headers))
        sheet.append([''] * len(headers))
        sheet.append(headers)
        
        header_border = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )
        
        for header_cell in sheet[5]:
            header_cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
            header_cell.font = Font(bold=True, color='FFFFFF', size=14)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = header_border
        
        row_index = 3
        for i in queryset:
            row = [
                i.product.name,
                i.cantidad_enviada,
                i.cantidad_vendida,
                i.cantidad_retirada,
                i.start_date,
                f"{i.total_deuda} $",
                "Si" if i.paid == True
                else "No",
                
            ]
            sheet.append(row)
            row_index += 1
        
        data_border = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )
        
        for row in sheet.iter_rows(min_row=6, max_row=row_index+2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = data_border
        
        # TOTAL CHARGED
        charged = []
        for i in queryset:
            if i.paid == True:
                charged.append(i.total_deuda)
        total_charged = sum(charged)
        sheet.merge_cells(start_row=5, start_column=11, end_row=5, end_column=13)
        merged_cell_total_charged_title = sheet.cell(row=5, column=11)
        merged_cell_total_charged_title.value = "TOTAL FACTURADO"
        merged_cell_total_charged_title.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell_total_charged_title.font = Font(bold=True, size=15, color='FFFFFF')
        sheet.merge_cells(start_row=6, start_column=11, end_row=6, end_column=13)
        merged_cell_total_charged = sheet.cell(row=6, column=11)
        merged_cell_total_charged.value = f"{total_charged} $"
        merged_cell_total_charged.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell_total_charged.font = Font(bold=True, size=15, color='000000')
        fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        merged_cell_total_charged_title.fill = fill
        merged_cell_total_charged.column = 11
        for row in sheet.iter_rows(min_row=5, max_row=6, min_col=11, max_col=13):
            for cell in row:
                cell.border = data_border
        filename = f"Reporte Consignas cliente {client_name[0]}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

class AddClientData(LoginRequiredMixin, CreateView):
    model = ClientData
    template_name = 'add_client_data.html'
    form_class = ClientDataForm

    def get_initial(self):
        initial = super().get_initial()
        client_id = self.kwargs['pk']
        initial['client'] = client_id
        initial['cantidad_enviada'] = 0
        initial['cantidad_vendida'] = 0
        initial['cantidad_retirada'] = 0
        initial['total_deuda'] = 0.00
        return initial

    def form_valid(self, form):
        product_id = form.cleaned_data['product'].id
        product = Product.objects.get(id=product_id)
        cantidad_enviada = form.cleaned_data['cantidad_enviada']
        cantidad_vendida = form.cleaned_data['cantidad_vendida']
        cantidad_retirada = form.cleaned_data['cantidad_retirada']

        total_deuda = Decimal(product.price) * Decimal(cantidad_enviada)
        form.instance.total_deuda = total_deuda

        

        product_quantity = product.quantity

        new_quantity = product_quantity - cantidad_enviada

        if cantidad_vendida > 0:
            new_quantity = product_quantity - cantidad_enviada + cantidad_vendida
        elif cantidad_retirada > 0:
            new_quantity = product_quantity - cantidad_enviada + cantidad_vendida + cantidad_retirada
        

        product.quantity = new_quantity
        product.save()

        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('clientsdata:ClientData', kwargs={'pk': self.kwargs['pk']})




class UpdateClientData(LoginRequiredMixin, UpdateView):
    model = ClientData
    template_name = 'update_client_data.html'
    form_class = ClientDataForm

    def form_valid(self, form):
        current_data = ClientData.objects.get(pk=self.kwargs['pk'])
        product_id = current_data.product.pk
        product = Product.objects.get(id=product_id)

        current_product_quantity = product.quantity

        last_cantidad_enviada = current_data.cantidad_enviada
        last_cantidad_vendida = current_data.cantidad_vendida
        last_cantidad_retirada = current_data.cantidad_retirada

        new_cantidad_enviada = form.cleaned_data['cantidad_enviada']
        new_cantidad_vendida = form.cleaned_data['cantidad_vendida']
        new_cantidad_retirada = form.cleaned_data['cantidad_retirada']
        paid = form.cleaned_data['paid']

        # Calcula total_deuda en función de la lógica de tu aplicación
        if paid:
            total_deuda = Decimal(product.price) * Decimal(new_cantidad_enviada - new_cantidad_vendida)
        else:
            total_deuda = Decimal(product.price) * Decimal(new_cantidad_enviada - new_cantidad_vendida)

        if last_cantidad_enviada < new_cantidad_enviada:
            new_product_quantity = current_product_quantity - new_cantidad_enviada
            product.quantity = new_product_quantity
            product.save()
        elif last_cantidad_enviada > new_cantidad_enviada:
            new_product_quantity = current_product_quantity + (last_cantidad_enviada - new_cantidad_enviada)
            product.quantity = new_product_quantity
            product.save()

        if last_cantidad_retirada < new_cantidad_retirada:
            new_product_quantity = current_product_quantity + (new_cantidad_retirada - last_cantidad_retirada)
            product.quantity = new_product_quantity
            product.save()
        elif last_cantidad_retirada > new_cantidad_retirada:
            new_product_quantity = current_product_quantity + new_cantidad_retirada
            product.quantity = new_product_quantity
            product.save()
        elif new_cantidad_retirada == last_cantidad_enviada:
            new_product_quantity = current_product_quantity + last_cantidad_enviada
            product.quantity = new_product_quantity
            product.save()

        # Asigna total_deuda al formulario
        form.instance.total_deuda = total_deuda

        # Guarda el formulario
        form.save()

        return super().form_valid(form)

    def get_success_url(self):
        client_id = self.object.client.id
        return reverse_lazy('clientsdata:ClientData', kwargs={'pk': client_id})



class DeleteClientData(LoginRequiredMixin, DeleteView):
    model = ClientData
    template_name = 'delete_client_data.html'
    success_url = reverse_lazy('clientsdata:ClientsData')

    

    def get_success_url(self):
        client_id = self.object.client.id
        return reverse_lazy('clientsdata:ClientData', kwargs={'pk': client_id})



#TOTAL CLIENTS DATA VIEWS
#########################


class ClientsDataList(LoginRequiredMixin, ListView):
    model = ClientData
    template_name = 'clients_data_list.html'
    context_object_name = 'clientsdata'

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        year_filter = self.request.GET.get('start_date')


        if search_query:
            queryset = queryset.filter(
                Q(client__id__icontains=search_query) |
                Q(product__name__icontains=search_query) |
                Q(cantidad_enviada__icontains=search_query) |
                Q(cantidad_vendida__icontains=search_query) |
                Q(cantidad_retirada__icontains=search_query) |
                Q(total_deuda__icontains=search_query)
                )
        
        if year_filter:
            queryset = queryset.filter(start_date__year=year_filter)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        all_clientsdata = ClientData.objects.all()
        context['all_clientsdata'] = len(all_clientsdata)

        
        years = ClientData.objects.dates('start_date', 'year', order='DESC')
        years = [year.year for year in years]
        context['selected_years'] = years
        
        # Obtener el año seleccionado en el filtro
        selected_year = self.request.GET.get('year')
        context['selected_year'] = selected_year

        return context
    


    def generate_excel_report_allclients(request):
        selected_year = request.GET.get('year')
        if selected_year is not None and selected_year != '':
            try:
                # Intenta convertir selected_year a entero
                selected_year = int(selected_year)
                
            except ValueError:
                print (selected_year)
                pass

            # Filtrar tu consulta solo si selected_year es un valor numérico
            queryset = ClientData.objects.filter(start_date__year=selected_year)
        else:
            queryset = ClientData.objects.all()

    
    
        workbook = Workbook()
        sheet = workbook.active
        
        sheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=8)
        merged_cell = sheet.cell(row=1, column=1)
        merged_cell.value = "Reporte total de consignas"
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.font = Font(bold=True, size=20, color='FFFFFF')
        fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        merged_cell.fill = fill
        sheet['A1'].font = Font(bold=True, size=20, color='FFFFFF', italic=True)
        row_dimension = sheet.row_dimensions[5]
        row_dimension.height = 30
        
        all_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'O', 'P']
        for column in all_columns:
            sheet.column_dimensions[column].width = 20
        
        headers = ['Cliente', 'Producto', 'Cantidad Enviada', 'Cantidad Vendida', 'Cantidad Retirada', 'Fecha', 'Total Deuda', 'Pagada']
        sheet.append([''] * len(headers))
        sheet.append([''] * len(headers))
        sheet.append([''] * len(headers))
        sheet.append(headers)
        
        header_border = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )
        
        for header_cell in sheet[5]:
            header_cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
            header_cell.font = Font(bold=True, color='FFFFFF', size=14)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = header_border
        
        row_index = 3
        for i in queryset:
            row = [
                i.client.fullname,
                i.product.name,
                i.cantidad_enviada,
                i.cantidad_vendida,
                i.cantidad_retirada,
                i.start_date,
                f"{i.total_deuda} $",
                "Si" if i.paid == True
                else "No",
                
            ]
            sheet.append(row)
            row_index += 1
        
        data_border = Border(
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
            left=Side(border_style="thin"),
            right=Side(border_style="thin")
        )
        
        for row in sheet.iter_rows(min_row=6, max_row=row_index+2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = data_border
        
        # TOTAL CHARGED
        charged = []
        for i in queryset:
            if i.paid == True:
                charged.append(i.total_deuda)
        total_charged = sum(charged)
        sheet.merge_cells(start_row=5, start_column=11, end_row=5, end_column=13)
        merged_cell_total_charged_title = sheet.cell(row=5, column=11)
        merged_cell_total_charged_title.value = "TOTAL FACTURADO"
        merged_cell_total_charged_title.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell_total_charged_title.font = Font(bold=True, size=15, color='FFFFFF')
        sheet.merge_cells(start_row=6, start_column=11, end_row=6, end_column=13)
        merged_cell_total_charged = sheet.cell(row=6, column=11)
        merged_cell_total_charged.value = f"{total_charged} $"
        merged_cell_total_charged.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell_total_charged.font = Font(bold=True, size=15, color='000000')
        fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        merged_cell_total_charged_title.fill = fill
        merged_cell_total_charged.column = 11
        for row in sheet.iter_rows(min_row=5, max_row=6, min_col=11, max_col=13):
            for cell in row:
                cell.border = data_border
        filename = "Reporte total de consignas.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response